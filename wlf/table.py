# -*- coding=UTF-8 -*-
"""Data table"""
from __future__ import (absolute_import, division, print_function,
                        unicode_literals)

import logging
import re

from openpyxl import Workbook
from six import text_type
from six.moves import range

from .progress import progress

LOGGER = logging.getLogger('com.wlf.table')


class RowTable(object):
    """Table based on row.  """

    l10n_dict = {}

    def __init__(self, rows=None, header=None):
        self.header = list(NestedData(header).to_tuple()) or []
        self.rows = []
        if rows:
            for row in rows:
                self.append(row)

    def append(self, row):
        """Append row to table.  """

        assert isinstance(row, dict), row
        row = NestedData(row).to_dict()
        for key in row:
            if key not in self.header:
                self.header.append(key)
        self.rows.append(row)

    def to_html(self, filename):
        # TODO
        pass

    def to_xlsx(self, filename):
        """Dump table to xlsx sheet.  """

        book = Workbook()
        sheet = book.active
        sheet.page_setup.fitToWidth = 1
        header = NestedData(self.header)
        header_rows = header.to_rows()

        # Create header
        for row_index, row in enumerate(header_rows):
            sheet.append(self.l10n(row))
            for column_index, item in enumerate(row):
                if row_index:
                    # Vertical merge
                    if item is None:
                        sheet.merge_cells(start_row=row_index - 1 + 1,
                                          end_row=row_index + 1,
                                          start_column=column_index + 1,
                                          end_column=column_index + 1)
                if column_index:
                    # Horizontal merge
                    prev_item = header_rows[row_index][column_index - 1]
                    if item == prev_item:
                        sheet.merge_cells(start_row=row_index + 1,
                                          end_row=row_index + 1,
                                          start_column=column_index - 1 + 1,
                                          end_column=column_index + 1)

        sheet.freeze_panes = sheet.cell(row=len(header_rows) + 1, column=1)

        # Dump data
        LOGGER.debug('Header: %s', self.header)
        for row in progress(self.rows, '导出表格 {}'.format(filename)):
            sheet_row = [row.get(i) for i in self.header]
            sheet.append(self.l10n(sheet_row))
        book.save(filename)
        LOGGER.info('导出表格: %s', filename)

    def l10n(self, text):
        """Localization.  """

        if not text:
            return text

        try:
            if isinstance(text, list):
                return [self.l10n(i) for i in text]

            for pattern, repl in self.l10n_dict.items():
                text = re.sub(text_type(pattern),
                              text_type(repl), text_type(text))
        except TypeError:
            LOGGER.warning('L10n fail:%s', text)

        return text


class Row(list):
    """Row in RowTable.  """
    pass


class NestedData(object):
    """List or dict nested sturcture. """

    def __init__(self, item):
        self.item = item

    def to_columns(self):
        """Convet to columns.  """

        columns = []
        item = self.item
        if isinstance(item, (list, tuple))\
                and len(item) == 2 \
                and isinstance(item[0], (str, text_type)) and isinstance(item[1], (tuple, list)):
            item = {item[0]: item[1]}
        if isinstance(item, (list, tuple)):
            for i in item:
                columns.extend(NestedData(i).to_columns())
        elif isinstance(item, dict):
            for k, v in item.items():
                for i in NestedData(v).to_columns():
                    column = [k]
                    column.extend(i)
                    columns.append(column)
        else:
            columns.append([item])
        return columns

    def to_rows(self):
        """Convet to rows.  """

        columns = self.to_columns()
        row_count = max(len(i) for i in columns)
        column_count = len(columns)
        rows = []
        _ = [rows.append([None] * column_count) for _ in range(row_count)]

        for column_index, column in enumerate(columns):
            for row_index, item in enumerate(column):
                rows[row_index][column_index] = item

        return rows

    def to_dict(self):
        """Convert to one-demension dict. """

        ret = {}
        columns = self.to_columns()
        for i in columns:
            if len(i) > 1:
                key = None
                for j in i[:-1][::-1]:
                    if key is None:
                        key = (j,)
                    else:
                        key = (j, key)
                ret[key] = i[-1]
        return ret

    def to_tuple(self):
        """To nested tuple.   """

        columns = self.to_columns()
        ret = []
        for i in columns:
            if len(i) > 1:
                key = None
                for j in i[::-1]:
                    if key is None:
                        key = (j,)
                    else:
                        key = (j, key)
                ret.append(key)
        return tuple(ret)
